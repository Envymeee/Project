from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('C:\WED2.xlsx', data_only=True) #Full directory better if less ang way
ws = wb.active #needs to be opened like active
import numpy as np
from ortools.linear_solver import pywraplp

GradeFirstVentilation =[]
GradeFirstIllumination =[]
GradeFirstNoise =[]
GradeFirstTemperature =[]
AvailableClassroomGradeFirst=0
AveragePerClassroom=1
x =[] #use as quick rename tool dont forget to change back to x

#RENAME TEMPLATE for temperature/illumination
"""
#Get Grade # ### from excel to data
for row in range(2, 11): #(From row 2 to row 10 read)
    for col in range(3, 4): #(From column 3 only)
        char = get_column_letter(col)
        x.append(ws[char + str(row)].value)


#Convert Grade # ### to 1 and 0
for i in range(len(x)):
    if x[i] >= 68: #First condition
        if x[i] <= 74: #Second condition delete for vent/noise
            x[i] = 1
    else:
        x[i] = 0
print(x)
"""
###################################################################################################################################################################################


#Get Grade 1 Ventilation 
for row in range(2, 11): #(From row 2 to row 10 read)
    for col in range(3, 4): #(From column 3 only)
        char = get_column_letter(col)
        GradeFirstVentilation.append(ws[char + str(row)].value)


#Convert Grade 1 Ventilation to 1 and 0
for i in range(len(GradeFirstVentilation)):
    if GradeFirstVentilation[i] >= 6:
        GradeFirstVentilation[i] = 1
    else:
        GradeFirstVentilation[i] = 0
print("Grade 1 Ventilation")
print(GradeFirstVentilation)

#Get Grade 1 Illumination
for row in range(2, 11): #(From row 2 to row 10 read)
    for col in range(5, 6): #(From column 3 only)
        char = get_column_letter(col)
        GradeFirstIllumination.append(ws[char + str(row)].value)

#Convert Grade 1 Ilumination
for i in range(len(GradeFirstIllumination)):
    if GradeFirstIllumination[i] <= 300:
        if GradeFirstIllumination[i] >=600:
            GradeFirstIllumination[i] = 1
        else:
            GradeFirstIllumination[i] = 0
    else:
        
        GradeFirstIllumination[i] = 0

print("Grade 1 Illumination")
print(GradeFirstIllumination)

#GET Grade 1 Noise 
for row in range(2, 11): #(From row 2 to row 10 read)
    for col in range(7, 8): #(From column 3 only)
        char = get_column_letter(col)
        GradeFirstNoise.append(ws[char + str(row)].value)


#Convert Grade 1 Noise to 1 and 0
for i in range(len(GradeFirstNoise)):
    if GradeFirstNoise[i] <= 35:
        GradeFirstNoise[i] = 1
    else:
        GradeFirstNoise[i] = 0
print("Grade 1 Noise")
print(GradeFirstNoise)

#Get Grade 1 Temperature
for row in range(2, 11): #(From row 2 to row 10 read)
    for col in range(3, 4): #(From column 3 only)
        char = get_column_letter(col)
        GradeFirstTemperature.append(ws[char + str(row)].value)


#Convert Grade # ### to 1 and 0
for i in range(len(GradeFirstTemperature)):
    if GradeFirstTemperature[i] >= 68: 
        if GradeFirstTemperature <=74:
            GradeFirstTemperature[i] = 1
    else:
        GradeFirstTemperature[i] = 0
print("Grade 1 Temperature")
print(GradeFirstTemperature)


###################################################################################################################################################################################

###Objective 1 Get the Good classrooms

GradeFirstResultFirstPart = []
GradeFirstResultFirstPart = [a * b for a, b in zip(GradeFirstVentilation,GradeFirstIllumination)] #Just replace the 2 variables to multiply
#print(GradeFirstWorkResultFirstPart)

GradeFirstResultSecondPart = []
GradeFirstResultSecondPart = [a * b for a, b in zip(GradeFirstNoise,GradeFirstTemperature)] #Just replace the 2 variables to multiply

GradeFirstTotal =[]
GradeFirstTotal = [a * b for a, b in zip(GradeFirstResultFirstPart,GradeFirstResultSecondPart)] #Just replace the 2 variables to multiply
print("GRADE 1 TOTAL")
print(GradeFirstTotal) #list version


print("GRADE 1 TOTAL in number") #Number version as converted
GradeFirstTotal = sum(GradeFirstTotal)
print(GradeFirstTotal)

###################################################################################################################################################################################

#Objective 2 just 
"""
(enrolled)
----------
(Good Classroom)
"""
GradeFirstEnrolled=ws['M3'].value #GET value from excel
print("Enrolled Grade 1 students")
print(GradeFirstEnrolled)

#TEST for 1 since lots of FAILED
GradeFirstTotal=1
print("Objective 2")
if GradeFirstTotal==0:
    print("Cant divide by 0")
else: 
    AvailableClassroomGradeFirst = GradeFirstEnrolled//GradeFirstTotal #will print out if Total>0
    AveragePerClassroom = GradeFirstEnrolled//AvailableClassroomGradeFirst


###################################################################################################################################################################################

#Objective 3 Bin packing program WIP


    
    
    

###################################################################################################################################################################################

#Objective 3 just math

ClassroomMax = 50


if AveragePerClassroom <= ClassroomMax:
    print(AveragePerClassroom)
if AveragePerClassroom > ClassroomMax:
    RemainderStudents= GradeFirstEnrolled%AvailableClassroomGradeFirst
    print("Excess students of "+RemainderStudents)
    AveragePerClassroom+=RemainderStudents
    print("New Average Per Classroom" +AveragePerClassroom)

